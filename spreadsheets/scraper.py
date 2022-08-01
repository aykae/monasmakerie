import json
import openpyxl


def main():
    priceScraper()

def priceScraper():
    loc = 'pricing.xlsx'
    f = openpyxl.load_workbook(loc)
    
    wax = f["Sheet1"]
    facial = f["Sheet2"]
    devices = f["Sheet3"]
    makeup = f["Sheet4"]

    data = {}

    waxScraper(data, wax)
    facialScraper(data, facial)
    deviceScraper(data, devices)
    makeupScraper(data, makeup)
    
    with open('../_data/pricing.json', 'w') as outfile:
        json.dump(data, outfile)


def waxScraper(data, sheet):
    data['wax'] = []
    for row in sheet.iter_rows(min_row=2):
        if row[1].value is None:
            break
        data['wax'].append({
            'name' : row[1].value,
            'time' : row[4].value,
            'price' : row[3].value
        })

def facialScraper(data, sheet):
    data['facials'] = []
    for row in sheet.iter_rows(min_row=2):
        data['facials'].append({
            'name' : row[0].value,
            'description' : row[5].value,
            'time' : row[4].value,
            'price' : row[3].value
        })

def deviceScraper(data, sheet):
    data['devices'] = []
    for row in sheet.iter_rows(min_row=2):
        if row[1].value is None:
            break
        data['devices'].append({
            'name' : row[1].value,
            'description' : row[5].value,
            'time' : row[4].value,
            'price' : row[3].value
        })

def makeupScraper(data, sheet):
    data['makeup'] = []
    for row in sheet.iter_rows(min_row=2):
        if row[1].value is None:
            break
        data['makeup'].append({
            'name' : row[1].value,
            'time' : row[4].value,
            'price' : row[3].value
        })


main()
